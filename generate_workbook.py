#!/usr/bin/env python3
"""Generate the NSC C3 Action Workbook (nsc-c3-workbook.xlsx)."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# ── Brand palette ──
CREAM      = "F8F1F2"
OFF_WHITE  = "F9F0F5"
PLUM       = "493751"
SAGE       = "A6C9BB"
PERIWINKLE = "A7AFC8"
LAVENDER   = "E0DAE7"
MAUVE      = "B375A0"
NAVY       = "3B3B58"
PEACH      = "E8B4A0"
WHITE_TEXT = "FFFFFF"

# ── Reusable fills ──
fill_cream      = PatternFill("solid", fgColor=CREAM)
fill_off_white  = PatternFill("solid", fgColor=OFF_WHITE)
fill_plum       = PatternFill("solid", fgColor=PLUM)
fill_sage       = PatternFill("solid", fgColor=SAGE)
fill_periwinkle = PatternFill("solid", fgColor=PERIWINKLE)
fill_lavender   = PatternFill("solid", fgColor=LAVENDER)
fill_mauve      = PatternFill("solid", fgColor=MAUVE)
fill_peach      = PatternFill("solid", fgColor=PEACH)

# ── Fonts ──
font_header      = Font(name="Cormorant Garamond", size=22, bold=True, color=WHITE_TEXT)
font_subheader   = Font(name="Cormorant Garamond", size=16, bold=True, color=WHITE_TEXT)
font_section      = Font(name="Calibri", size=13, bold=True, color=WHITE_TEXT)
font_section_label = Font(name="Calibri", size=12, bold=True, color=PLUM)
font_body        = Font(name="Calibri", size=11, color=NAVY)
font_body_bold   = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_instruction = Font(name="Calibri", size=10, italic=True, color=PERIWINKLE)
font_total_white = Font(name="Calibri", size=11, bold=True, color=WHITE_TEXT)
font_total_navy  = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_cover_desc  = Font(name="Calibri", size=12, color=NAVY)
font_cover_nav   = Font(name="Calibri", size=11, color=NAVY)
font_cover_nav_bold = Font(name="Calibri", size=11, bold=True, color=PLUM)

# ── Borders ──
thin_pw = Side(style="thin", color=PERIWINKLE)
border_input = Border(top=thin_pw, bottom=thin_pw, left=thin_pw, right=thin_pw)
border_none = Border()

# ── Alignments ──
align_left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_right  = Alignment(horizontal="right", vertical="center")

# ── Protection ──
prot_locked   = Protection(locked=True)
prot_unlocked = Protection(locked=False)

# ── Tab colors ──
TAB_COLORS = [PLUM, MAUVE, SAGE, LAVENDER, PERIWINKLE, SAGE, MAUVE]

# ── Helpers ──

def style_cell(cell, font=None, fill=None, border=None, alignment=None, protection=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment
    if protection: cell.protection = protection
    if number_format: cell.number_format = number_format


def section_header(ws, row, col_start, col_end, text):
    """Plum fill, white text section header merged across columns."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    style_cell(cell, font=font_section, fill=fill_plum, alignment=align_left, protection=prot_locked)
    for c in range(col_start, col_end + 1):
        style_cell(ws.cell(row=row, column=c), fill=fill_plum, protection=prot_locked)
    ws.row_dimensions[row].height = 28
    return row + 1


def section_label(ws, row, col_start, col_end, text):
    """Sage fill, plum text sub-header."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    style_cell(cell, font=font_section_label, fill=fill_sage, alignment=align_left, protection=prot_locked)
    for c in range(col_start, col_end + 1):
        style_cell(ws.cell(row=row, column=c), fill=fill_sage, protection=prot_locked)
    ws.row_dimensions[row].height = 24
    return row + 1


def instruction_row(ws, row, col_start, col_end, text):
    """Cream fill, periwinkle italic instruction text."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    style_cell(cell, font=font_instruction, fill=fill_cream, alignment=align_left, protection=prot_locked)
    for c in range(col_start, col_end + 1):
        style_cell(ws.cell(row=row, column=c), fill=fill_cream, protection=prot_locked)
    return row + 1


def table_header_row(ws, row, headers, col_start=1):
    """Plum fill column headers."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        style_cell(cell, font=font_section, fill=fill_plum, alignment=align_center, protection=prot_locked, border=border_input)
    ws.row_dimensions[row].height = 24
    return row + 1


def input_row(ws, row, col_start, col_end, defaults=None):
    """Off-white input cells."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        val = defaults.get(c - col_start) if defaults else None
        if val is not None:
            cell.value = val
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_left, protection=prot_unlocked)
    return row + 1


def amount_input_row(ws, row, col_start, col_end, amount_cols=None, defaults=None):
    """Input row where specific columns are formatted as currency."""
    amount_cols = amount_cols or []
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        val = defaults.get(c - col_start) if defaults else None
        if val is not None:
            cell.value = val
        fmt = '$#,##0.00' if (c - col_start) in amount_cols else None
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_right if (c - col_start) in amount_cols else align_left,
                   protection=prot_unlocked, number_format=fmt)
    return row + 1


def formula_cell(ws, row, col, formula, primary=False):
    """Lavender or mauve formula cell."""
    cell = ws.cell(row=row, column=col, value=formula)
    if primary:
        style_cell(cell, font=font_total_white, fill=fill_mauve, border=border_input,
                   alignment=align_right, protection=prot_locked, number_format='$#,##0.00')
    else:
        style_cell(cell, font=font_total_navy, fill=fill_lavender, border=border_input,
                   alignment=align_right, protection=prot_locked, number_format='$#,##0.00')
    return cell


def summary_label(ws, row, col, text, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    style_cell(cell, font=font_body_bold, fill=fill_lavender, border=border_input,
               alignment=align_left, protection=prot_locked)
    for c in range(col, col + span):
        style_cell(ws.cell(row=row, column=c), fill=fill_lavender, border=border_input, protection=prot_locked)


def set_col_widths(ws, widths):
    """widths: dict of col_index (1-based) -> width in chars."""
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def fill_background(ws, max_row, max_col):
    """Fill all empty cells with cream background."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill == PatternFill() or cell.fill.fgColor is None or cell.fill.fgColor.rgb == '00000000':
                cell.fill = fill_cream


def apply_yn_validation(ws, col, row_start, row_end):
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.error = "Please enter Y or N"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    for r in range(row_start, row_end + 1):
        dv.add(ws.cell(row=r, column=col))


def apply_c3_validation(ws, col, row_start, row_end):
    dv = DataValidation(type="list", formula1='"Keep,Cut,Cancel,Combine"', allow_blank=True)
    dv.error = "Please select Keep, Cut, Cancel, or Combine"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    for r in range(row_start, row_end + 1):
        dv.add(ws.cell(row=r, column=col))


# ══════════════════════════════════════════════════════════════
# WORKBOOK CREATION
# ══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# Remove default sheet
default_sheet = wb.active
default_sheet.title = "Cover"

tab_names = [
    "Cover",
    "Numbers Snapshot",
    "C3 First Look",
    "Cut List",
    "Cancel List",
    "Combine List",
    "C3 Action Plan",
]

sheets = {}
for i, name in enumerate(tab_names):
    if i == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = TAB_COLORS[i]
    sheets[name] = ws

# ══════════════════════════════════════════════════════════════
# TAB 0: COVER
# ══════════════════════════════════════════════════════════════

ws = sheets["Cover"]
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
set_col_widths(ws, {1: 5, 2: 60, 3: 5})

r = 2
ws.merge_cells("B2:B2")
cell = ws.cell(row=r, column=2, value="Spending Is Data: A 30-Day Pattern Recognition Course")
style_cell(cell, font=Font(name="Cormorant Garamond", size=24, bold=True, color=PLUM),
           fill=fill_cream, alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
ws.row_dimensions[r].height = 40

r = 4
cell = ws.cell(row=r, column=2, value="C3 Action Workbook")
style_cell(cell, font=Font(name="Cormorant Garamond", size=20, bold=True, color=MAUVE),
           fill=fill_cream, alignment=align_left)
ws.row_dimensions[r].height = 34

r = 6
cell = ws.cell(row=r, column=2,
               value="Six worksheets that build your personal Cut, Cancel, and Combine action plan.")
style_cell(cell, font=font_cover_desc, fill=fill_cream, alignment=align_left)
ws.row_dimensions[r].height = 24

r = 8
cell = ws.cell(row=r, column=2, value="Tab Navigation Guide")
style_cell(cell, font=Font(name="Cormorant Garamond", size=16, bold=True, color=PLUM),
           fill=fill_sage, alignment=align_left)
for c in range(1, 4):
    style_cell(ws.cell(row=r, column=c), fill=fill_sage)
ws.row_dimensions[r].height = 28

nav_items = [
    ("Numbers Snapshot", "Capture your income, expenses, and subscriptions at a glance."),
    ("C3 First Look", "Tag each subscription with your first instinct: Keep, Cut, Cancel, or Combine."),
    ("Cut List", "Document spending habits you want to reduce in frequency."),
    ("Cancel List", "List subscriptions and services to stop or downgrade."),
    ("Combine List", "Find consolidation, sharing, and replacement opportunities."),
    ("C3 Action Plan", "Pull all three strategies together into one sequenced plan."),
]

for i, (tab, desc) in enumerate(nav_items):
    r = 10 + i * 2
    cell = ws.cell(row=r, column=2, value=f"{tab}  —  {desc}")
    style_cell(cell, font=font_cover_nav, fill=fill_cream, alignment=align_left)
    ws.row_dimensions[r].height = 22

fill_background(ws, 22, 3)
ws.print_area = "A1:C22"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# TAB 1: NUMBERS SNAPSHOT
# ══════════════════════════════════════════════════════════════

ws = sheets["Numbers Snapshot"]
set_col_widths(ws, {1: 30, 2: 18, 3: 18, 4: 5})
ws.freeze_panes = "A2"

r = 1
r = section_header(ws, r, 1, 3, "Numbers Snapshot")

# ── Section 1: Income ──
r = section_label(ws, r, 1, 3, "Section 1 — Income")
r = instruction_row(ws, r, 1, 3, "Enter each income source and its monthly amount.")
r = table_header_row(ws, r, ["Income Source", "Monthly Amount"], col_start=1)
income_data_start = r
for i in range(5):
    r = amount_input_row(ws, r, 1, 2, amount_cols=[1])
income_data_end = r - 1

# Total income
summary_label(ws, r, 1, "Total Income")
formula_cell(ws, r, 2, f"=SUM(B{income_data_start}:B{income_data_end})", primary=True)
r += 1
r += 1  # blank row

# ── Section 2: Fixed Expenses ──
r = section_label(ws, r, 1, 3, "Section 2 — Fixed Expenses")
r = instruction_row(ws, r, 1, 3, "List your fixed monthly expenses (rent, utilities, insurance, etc.).")
r = table_header_row(ws, r, ["Expense Name", "Monthly Amount"], col_start=1)
fixed_data_start = r
for i in range(12):
    r = amount_input_row(ws, r, 1, 2, amount_cols=[1])
fixed_data_end = r - 1

summary_label(ws, r, 1, "Total Fixed Expenses")
formula_cell(ws, r, 2, f"=SUM(B{fixed_data_start}:B{fixed_data_end})", primary=True)
fixed_total_row = r
r += 1
r += 1

# ── Section 3: Variable Spending ──
r = section_label(ws, r, 1, 3, "Section 3 — Variable Spending")
r = instruction_row(ws, r, 1, 3, "Enter your typical monthly spend for each category.")
r = table_header_row(ws, r, ["Category", "Monthly Amount"], col_start=1)
var_labels = ["Groceries", "Dining out / takeout", "Entertainment",
              "Shopping (clothing + household)", "Personal care", "Subscriptions", "Other"]
var_data_start = r
for label in var_labels:
    r = amount_input_row(ws, r, 1, 2, amount_cols=[1], defaults={0: label})
var_data_end = r - 1

summary_label(ws, r, 1, "Total Variable Spending")
formula_cell(ws, r, 2, f"=SUM(B{var_data_start}:B{var_data_end})", primary=True)
var_total_row = r
r += 1
r += 1

# ── Section 4: Subscription Audit ──
r = section_label(ws, r, 1, 3, "Section 4 — Subscription Audit")
r = instruction_row(ws, r, 1, 3, "List every subscription you pay for. Mark Y if you use it regularly, N if you don't.")
r = table_header_row(ws, r, ["Subscription Name", "Monthly Cost", "Use Regularly (Y/N)"], col_start=1)
sub_data_start = r
for i in range(15):
    row_r = r
    r = amount_input_row(ws, r, 1, 3, amount_cols=[1])
sub_data_end = r - 1
apply_yn_validation(ws, 3, sub_data_start, sub_data_end)

summary_label(ws, r, 1, "Total Monthly Subscriptions")
formula_cell(ws, r, 2, f"=SUM(B{sub_data_start}:B{sub_data_end})", primary=True)
sub_total_row = r
# callout cell
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=3)
cell = ws.cell(row=r, column=3)
cell.value = f'="Total monthly subscriptions: $"&TEXT(B{r},"#,##0.00")'
style_cell(cell, font=font_total_navy, fill=fill_lavender, border=border_input,
           alignment=align_left, protection=prot_locked)
r += 1
r += 1

# ── Section 5: Summary ──
r = section_label(ws, r, 1, 3, "Section 5 — Summary")
income_total = f"B{income_data_start - 1 + 5 + 1}"  # we need exact row refs
# Let me use the actual row numbers we tracked
income_total_row_num = income_data_end + 1

labels_formulas = [
    ("Total In", f"=B{income_total_row_num}"),
    ("Total Fixed", f"=B{fixed_total_row}"),
    ("Total Variable", f"=B{var_total_row}"),
    ("Total Subscriptions", f"=B{sub_total_row}"),
    ("Estimated Monthly Surplus/Gap", f"=B{income_total_row_num}-B{fixed_total_row}-B{var_total_row}"),
]
for label, formula in labels_formulas:
    summary_label(ws, r, 1, label)
    is_primary = "Surplus" in label
    formula_cell(ws, r, 2, formula, primary=is_primary)
    r += 1

fill_background(ws, r + 2, 4)
ws.print_area = f"A1:C{r+1}"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# TAB 2: C3 FIRST LOOK
# ══════════════════════════════════════════════════════════════

ws = sheets["C3 First Look"]
set_col_widths(ws, {1: 30, 2: 18, 3: 22, 4: 30, 5: 5})
ws.freeze_panes = "A2"

r = 1
r = section_header(ws, r, 1, 4, "C3 First Look")
r = instruction_row(ws, r, 1, 4,
    "This is a first-instinct pass — no commitments required. Tag each subscription with Keep, Cut, Cancel, or Combine.")

r = table_header_row(ws, r, ["Subscription Name", "Monthly Cost", "First Instinct", "Notes"], col_start=1)
c3fl_data_start = r
for i in range(15):
    # Subscription name pulls from Tab 1
    cell_name = ws.cell(row=r, column=1)
    cell_name.value = f"='Numbers Snapshot'!A{sub_data_start + i}"
    style_cell(cell_name, font=font_body, fill=fill_lavender, border=border_input,
               alignment=align_left, protection=prot_locked)

    cell_cost = ws.cell(row=r, column=2)
    cell_cost.value = f"='Numbers Snapshot'!B{sub_data_start + i}"
    style_cell(cell_cost, font=font_body, fill=fill_lavender, border=border_input,
               alignment=align_right, protection=prot_locked, number_format='$#,##0.00')

    # First Instinct - dropdown
    cell_fi = ws.cell(row=r, column=3)
    style_cell(cell_fi, font=font_body, fill=fill_off_white, border=border_input,
               alignment=align_center, protection=prot_unlocked)

    # Notes - free text
    cell_notes = ws.cell(row=r, column=4)
    style_cell(cell_notes, font=font_body, fill=fill_off_white, border=border_input,
               alignment=align_left, protection=prot_unlocked)
    r += 1
c3fl_data_end = r - 1

apply_c3_validation(ws, 3, c3fl_data_start, c3fl_data_end)

r += 1
# Summary cells
r = section_label(ws, r, 1, 4, "Summary")

summary_label(ws, r, 1, "Count of rows marked Cancel", span=2)
formula_cell(ws, r, 3, f'=COUNTIF(C{c3fl_data_start}:C{c3fl_data_end},"Cancel")')
ws.cell(row=r, column=3).number_format = '0'
r += 1

summary_label(ws, r, 1, "Total estimated monthly savings if all Cancel rows were acted on", span=2)
formula_cell(ws, r, 3, f'=SUMIF(C{c3fl_data_start}:C{c3fl_data_end},"Cancel",B{c3fl_data_start}:B{c3fl_data_end})')
r += 1

# Callout cell
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
cell = ws.cell(row=r, column=1)
cell.value = f'="If you canceled everything marked Cancel: $"&TEXT(SUMIF(C{c3fl_data_start}:C{c3fl_data_end},"Cancel",B{c3fl_data_start}:B{c3fl_data_end}),"#,##0.00")&"/month"'
style_cell(cell, font=font_total_white, fill=fill_mauve, border=border_input,
           alignment=align_center, protection=prot_locked)
for c in range(1, 5):
    style_cell(ws.cell(row=r, column=c), fill=fill_mauve, border=border_input, protection=prot_locked)

fill_background(ws, r + 2, 5)
ws.print_area = f"A1:D{r+1}"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# TAB 3: CUT LIST
# ══════════════════════════════════════════════════════════════

ws = sheets["Cut List"]
set_col_widths(ws, {1: 28, 2: 22, 3: 20, 4: 35, 5: 22, 6: 5})
ws.freeze_panes = "A2"

r = 1
r = section_header(ws, r, 1, 5, "Cut List")
r = instruction_row(ws, r, 1, 5,
    "Document spending habits to reduce in frequency, not eliminate entirely.")

r = table_header_row(ws, r, [
    "Spending Habit", "Current Frequency", "Cut Target",
    "How/When I Will Decide (day, time, or context)", "Estimated Monthly Savings"
], col_start=1)
cut_data_start = r
for i in range(10):
    r = amount_input_row(ws, r, 1, 5, amount_cols=[4])
cut_data_end = r - 1

r += 1
summary_label(ws, r, 1, "Total estimated monthly savings from cuts", span=4)
formula_cell(ws, r, 5, f"=SUM(E{cut_data_start}:E{cut_data_end})", primary=True)
cut_total_row = r
r += 1

# Summary cell
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cell = ws.cell(row=r, column=1)
cell.value = f'="Total estimated monthly savings from cuts: $"&TEXT(E{cut_total_row},"#,##0.00")'
style_cell(cell, font=font_total_white, fill=fill_mauve, border=border_input,
           alignment=align_center, protection=prot_locked)
for c in range(1, 6):
    style_cell(ws.cell(row=r, column=c), fill=fill_mauve, border=border_input, protection=prot_locked)
r += 1

# Open cells
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cell = ws.cell(row=r, column=1, value="Which Chasing Pattern shows up most in my Cut list:")
style_cell(cell, font=font_section_label, fill=fill_sage, alignment=align_left, protection=prot_locked)
for c in range(1, 6):
    style_cell(ws.cell(row=r, column=c), fill=fill_sage, protection=prot_locked)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cell = ws.cell(row=r, column=1)
style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
           alignment=align_left, protection=prot_unlocked)
for c in range(1, 6):
    style_cell(ws.cell(row=r, column=c), fill=fill_off_white, border=border_input, protection=prot_unlocked)
ws.row_dimensions[r].height = 40
r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cell = ws.cell(row=r, column=1, value="The one cut I am starting this week:")
style_cell(cell, font=font_section_label, fill=fill_sage, alignment=align_left, protection=prot_locked)
for c in range(1, 6):
    style_cell(ws.cell(row=r, column=c), fill=fill_sage, protection=prot_locked)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cell = ws.cell(row=r, column=1)
style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
           alignment=align_left, protection=prot_unlocked)
for c in range(1, 6):
    style_cell(ws.cell(row=r, column=c), fill=fill_off_white, border=border_input, protection=prot_unlocked)
ws.row_dimensions[r].height = 40

fill_background(ws, r + 2, 6)
ws.print_area = f"A1:E{r+1}"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# TAB 4: CANCEL LIST
# ══════════════════════════════════════════════════════════════

ws = sheets["Cancel List"]
set_col_widths(ws, {1: 28, 2: 18, 3: 25, 4: 22, 5: 5})
ws.freeze_panes = "A2"

r = 1
r = section_header(ws, r, 1, 4, "Cancel List")
r = instruction_row(ws, r, 1, 4,
    "Document subscriptions and services to stop entirely or downgrade.")

# ── Table 1: Full Cancels ──
r = section_label(ws, r, 1, 4, "Table 1 — Full Cancels")
r = table_header_row(ws, r, [
    "Subscription / Service", "Monthly Cost", "Last Time I Used It", "Target Cancel Date"
], col_start=1)
cancel_data_start = r
for i in range(10):
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        fmt = '$#,##0.00' if c == 2 else ('yyyy-mm-dd' if c in (3, 4) else None)
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_right if c == 2 else align_left,
                   protection=prot_unlocked, number_format=fmt)
    r += 1
cancel_data_end = r - 1

summary_label(ws, r, 1, "Total monthly recovery from full cancels", span=1)
formula_cell(ws, r, 2, f"=SUM(B{cancel_data_start}:B{cancel_data_end})", primary=True)
cancel_total_row = r
r += 1
r += 1

# ── Table 2: Downgrades ──
r = section_label(ws, r, 1, 4, "Table 2 — Downgrades")
r = table_header_row(ws, r, [
    "Subscription / Service", "Current Cost", "Free or Lower Tier (Y/N)", "Notes"
], col_start=1)
dg_data_start = r
for i in range(8):
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        fmt = '$#,##0.00' if c == 2 else None
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_right if c == 2 else align_left,
                   protection=prot_unlocked, number_format=fmt)
    r += 1
dg_data_end = r - 1
apply_yn_validation(ws, 3, dg_data_start, dg_data_end)

summary_label(ws, r, 1, "Total monthly reduction from downgrades", span=1)
formula_cell(ws, r, 2, f"=SUM(B{dg_data_start}:B{dg_data_end})", primary=True)
dg_total_row = r
r += 1
r += 1

# Combined
r = section_label(ws, r, 1, 4, "Combined Summary")
summary_label(ws, r, 1, "Combined monthly recovery", span=1)
formula_cell(ws, r, 2, f"=B{cancel_total_row}+B{dg_total_row}", primary=True)
cancel_combined_row = r
r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
cell = ws.cell(row=r, column=1, value="The cancel I keep putting off, and why:")
style_cell(cell, font=font_section_label, fill=fill_sage, alignment=align_left, protection=prot_locked)
for c in range(1, 5):
    style_cell(ws.cell(row=r, column=c), fill=fill_sage, protection=prot_locked)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
cell = ws.cell(row=r, column=1)
style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
           alignment=align_left, protection=prot_unlocked)
for c in range(1, 5):
    style_cell(ws.cell(row=r, column=c), fill=fill_off_white, border=border_input, protection=prot_unlocked)
ws.row_dimensions[r].height = 40

fill_background(ws, r + 2, 5)
ws.print_area = f"A1:D{r+1}"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# TAB 5: COMBINE LIST
# ══════════════════════════════════════════════════════════════

ws = sheets["Combine List"]
set_col_widths(ws, {1: 30, 2: 30, 3: 22, 4: 5})
ws.freeze_panes = "A2"

r = 1
r = section_header(ws, r, 1, 3, "Combine List")
r = instruction_row(ws, r, 1, 3,
    "Document consolidation, sharing, and replacement opportunities.")

# ── Table 1: Consolidation ──
r = section_label(ws, r, 1, 3, "Table 1 — Consolidation")
r = table_header_row(ws, r, ["Current Setup (multiple tools)", "What to Consolidate Into", "Monthly Savings"], col_start=1)
consol_data_start = r
for i in range(6):
    r = amount_input_row(ws, r, 1, 3, amount_cols=[2])
consol_data_end = r - 1

summary_label(ws, r, 1, "Total Consolidation Savings", span=2)
formula_cell(ws, r, 3, f"=SUM(C{consol_data_start}:C{consol_data_end})")
consol_total_row = r
r += 1
r += 1

# ── Table 2: Sharing ──
r = section_label(ws, r, 1, 3, "Table 2 — Sharing")
r = table_header_row(ws, r, ["Service / Subscription", "Could Share With", "Potential Monthly Savings"], col_start=1)
share_data_start = r
for i in range(6):
    r = amount_input_row(ws, r, 1, 3, amount_cols=[2])
share_data_end = r - 1

summary_label(ws, r, 1, "Total Sharing Savings", span=2)
formula_cell(ws, r, 3, f"=SUM(C{share_data_start}:C{share_data_end})")
share_total_row = r
r += 1
r += 1

# ── Table 3: Replacement ──
r = section_label(ws, r, 1, 3, "Table 3 — Replacement")
r = table_header_row(ws, r, ["Current Expensive Option", "Cheaper Alternative That Does the Same Job", "Monthly Savings"], col_start=1)
repl_data_start = r
for i in range(6):
    r = amount_input_row(ws, r, 1, 3, amount_cols=[2])
repl_data_end = r - 1

summary_label(ws, r, 1, "Total Replacement Savings", span=2)
formula_cell(ws, r, 3, f"=SUM(C{repl_data_start}:C{repl_data_end})")
repl_total_row = r
r += 1
r += 1

# Combined summary
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
cell = ws.cell(row=r, column=1)
cell.value = f'="Total estimated monthly savings from Combine strategies: $"&TEXT(C{consol_total_row}+C{share_total_row}+C{repl_total_row},"#,##0.00")'
style_cell(cell, font=font_total_white, fill=fill_mauve, border=border_input,
           alignment=align_center, protection=prot_locked)
for c in range(1, 4):
    style_cell(ws.cell(row=r, column=c), fill=fill_mauve, border=border_input, protection=prot_locked)
combine_grand_total_row = r
# We need a plain formula cell for Tab 6 to reference
# Put the numeric total in column 3 of the next row
r += 1
summary_label(ws, r, 1, "Total Combine Savings (numeric)", span=2)
formula_cell(ws, r, 3, f"=C{consol_total_row}+C{share_total_row}+C{repl_total_row}", primary=True)
combine_numeric_total_row = r

fill_background(ws, r + 2, 4)
ws.print_area = f"A1:C{r+1}"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# TAB 6: C3 ACTION PLAN
# ══════════════════════════════════════════════════════════════

ws = sheets["C3 Action Plan"]
set_col_widths(ws, {1: 30, 2: 30, 3: 30, 4: 5})
ws.freeze_panes = "A2"

r = 1
r = section_header(ws, r, 1, 3, "C3 Action Plan")
r = instruction_row(ws, r, 1, 3,
    "Your complete C3 strategy — all values pull automatically from your Cut, Cancel, and Combine tabs.")

# ── Section 1: The Math ──
r = section_label(ws, r, 1, 3, "Section 1 — The Math")

math_items = [
    ("Monthly savings from Cut", f"='Cut List'!E{cut_total_row}"),
    ("Monthly savings from Cancel", f"='Cancel List'!B{cancel_combined_row}"),
    ("Monthly savings from Combine", f"='Combine List'!C{combine_numeric_total_row}"),
]
for label, formula in math_items:
    summary_label(ws, r, 1, label)
    formula_cell(ws, r, 2, formula)
    r += 1

# Total monthly recovery
summary_label(ws, r, 1, "Total monthly recovery")
formula_cell(ws, r, 2,
    f"='Cut List'!E{cut_total_row}+'Cancel List'!B{cancel_combined_row}+'Combine List'!C{combine_numeric_total_row}",
    primary=True)
total_monthly_row = r
r += 1

# Projected annual - mauve/lavender highlight
summary_label(ws, r, 1, "Projected annual recovery")
cell = formula_cell(ws, r, 2, f"=B{total_monthly_row}*12", primary=True)
# Extra emphasis with mauve
style_cell(cell, font=Font(name="Calibri", size=14, bold=True, color=WHITE_TEXT),
           fill=fill_mauve, border=border_input, alignment=align_right, protection=prot_locked,
           number_format='$#,##0.00')
ws.row_dimensions[r].height = 30
r += 1
r += 1

# ── Section 2: Pattern to Strategy ──
r = section_label(ws, r, 1, 3, "Section 2 — Pattern to Strategy")
r = table_header_row(ws, r, ["My Active Pattern", "C3 Strategy That Fits", "What That Looks Like for Me"], col_start=1)
for i in range(4):
    r = input_row(ws, r, 1, 3)
r += 1

# ── Section 3: Action Sequence ──
r = section_label(ws, r, 1, 3, "Section 3 — Action Sequence")

# This Week
r = section_label(ws, r, 1, 2, "This Week (Days 15 to 21)")
r_label = r  # skip header placement
r = table_header_row(ws, r, ["Item", "Date"], col_start=1)
for i in range(5):
    for c in range(1, 3):
        cell = ws.cell(row=r, column=c)
        fmt = 'yyyy-mm-dd' if c == 2 else None
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_left, protection=prot_unlocked, number_format=fmt)
    r += 1
r += 1

# By End of Challenge
r = section_label(ws, r, 1, 2, "By End of Challenge (Day 30)")
r = table_header_row(ws, r, ["Item", "Date"], col_start=1)
for i in range(5):
    for c in range(1, 3):
        cell = ws.cell(row=r, column=c)
        fmt = 'yyyy-mm-dd' if c == 2 else None
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_left, protection=prot_unlocked, number_format=fmt)
    r += 1
r += 1

# After the Challenge
r = section_label(ws, r, 1, 2, "After the Challenge")
r = table_header_row(ws, r, ["Item", "Date"], col_start=1)
for i in range(5):
    for c in range(1, 3):
        cell = ws.cell(row=r, column=c)
        fmt = 'yyyy-mm-dd' if c == 2 else None
        style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
                   alignment=align_left, protection=prot_unlocked, number_format=fmt)
    r += 1
r += 1

# ── Section 4: One Sentence ──
r = section_label(ws, r, 1, 3, "Section 4 — One Sentence")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
cell = ws.cell(row=r, column=1, value="In one sentence, what I am changing and why:")
style_cell(cell, font=font_section_label, fill=fill_sage, alignment=align_left, protection=prot_locked)
for c in range(1, 4):
    style_cell(ws.cell(row=r, column=c), fill=fill_sage, protection=prot_locked)
r += 1

ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=3)
cell = ws.cell(row=r, column=1)
style_cell(cell, font=font_body, fill=fill_off_white, border=border_input,
           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
           protection=prot_unlocked)
for rr in range(r, r+2):
    for c in range(1, 4):
        style_cell(ws.cell(row=rr, column=c), fill=fill_off_white, border=border_input, protection=prot_unlocked)
ws.row_dimensions[r].height = 50

fill_background(ws, r + 3, 4)
ws.print_area = f"A1:C{r+2}"
ws.protection.sheet = True
ws.protection.enable()

# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════

output_path = "src/features/nsc-c3-workbook.xlsx"
wb.save(output_path)
print(f"Workbook saved to {output_path}")
print("Verifying formulas...")

# Verify no obvious formula errors by checking all cells
wb2 = openpyxl.load_workbook(output_path)
issues = []
for ws_name in wb2.sheetnames:
    ws = wb2[ws_name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Check for common formula issues
                val = cell.value
                if "#REF!" in val or "#DIV/0!" in val or "#VALUE!" in val or "#NAME?" in val:
                    issues.append(f"  {ws_name}!{cell.coordinate}: {val}")
                # Check cross-sheet refs have valid sheet names
                if "'" in val:
                    import re
                    refs = re.findall(r"'([^']+)'!", val)
                    for ref in refs:
                        if ref not in wb2.sheetnames:
                            issues.append(f"  {ws_name}!{cell.coordinate}: bad sheet ref '{ref}' in {val}")

if issues:
    print("FORMULA ISSUES FOUND:")
    for iss in issues:
        print(iss)
else:
    print("Zero formula errors detected.")

print(f"\nTabs: {wb2.sheetnames}")
print(f"Total sheets: {len(wb2.sheetnames)}")
